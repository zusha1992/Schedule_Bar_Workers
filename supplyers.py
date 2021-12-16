import math
import openpyxl as op
from openpyxl.workbook import Workbook
from openpyxl.styles import Font





supplyers = ['Hacerem', 'Israko', 'Ristreto', 'Flam', '123', 'Elitzur', 'Castel', 'Ayalon' ]

supply_items = ['Date', 'Item', 'units', 'Price', 'overall']

months = ["january", "February", "march", "April", "May", "june", "July", "August", "September", "October", "Novembre",
          "December"]




def add_sheet(wb, name, categories):
    ws = wb.create_sheet(name)
    for column, category in enumerate(categories):
        ws.cell(1, column + 1, category)

def creat_workbook(supplyers):
    wb = op.Workbook()
    for supp in supplyers:
        add_sheet(wb, supp, supply_items)
    wb.save("supplyers tracing.xlsx")
    # add_sheet(wb, months[month - 1] + '20' + str(year),)


def find_first_empty(ws, date=''):
    current_row = 1
    while ws.cell(current_row, 1).value and ws.cell(current_row, 1).value != '':
        current_row += 1
    return current_row


def update_supplyer(ws, date, ):
    empty_row = find_first_empty(ws, date)
    item = input('Please enter item name: ')
    units = input('Please enter units number: ')
    price = input('Please enter price: ')
    overall = int(price) * int(units)
    details = [date, item, units, price, overall]
    for columns, detail in enumerate(details):
        ws.cell(empty_row, columns + 1, detail)


def update_sum(ws):
    current_row = find_first_empty(ws)
    sum = 0
    for i in range(2, current_row):
        sum += int(ws.cell(i, 5).value)
    # ws.cell(current_row + 3, i+ 1, '').font = Font(bold=False)
    ws.cell(current_row + 4, 5, sum).font = Font(bold=True)


wb = op.load_workbook('supplyers tracing.xlsx')
# update_supplyer(wb['Hacerem'], '1.10.21')
update_sum(wb['Hacerem'])

wb.save("supplyers tracing.xlsx")



# creat_workbook(supplyers)
