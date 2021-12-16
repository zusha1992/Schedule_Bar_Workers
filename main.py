import math
import openpyxl as op
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

# open new workbook
# wb = op.Workbook()

# load workbook
# wb= op.load_workbook('worker tracing.xls')


categories = ["Date", "Hours", "Base salary", "Tip", "Overall", "Completion", "Additional hours", "tip_percentage",
              "Shabbat hours", "comments"]

worker_list = ["MichalD", "MichalH", "Yoad", "Alon", "Rinat", "Aime", "Nitai", "Gilad", "Noam", "Omri", "Niv"]

months = ["january", "February", "march", "April", "May", "june", "July", "August", "September", "October", "Novembre",
          "December"]

income_list = ['date', 'x', "credit", "cash", "tip", "overall income"]


def add_sheet(wb, name, categories):
    ws = wb.create_sheet(name)
    for column, category in enumerate(categories):
        ws.cell(1, column + 1, category)

def find_first_empty(ws, date=''):
    """
    finds a empty cell in a Worksheet
    :param ws: Worksheet object imported from openpyxl.
    :param date:
    :return: returns index of empty cell in Worksheet
    """
    current_row = 1
    while ws.cell(current_row, 1).value and ws.cell(current_row, 1).value != date:
        current_row += 1
    return current_row


def calculate_base_salary(hours, isPik, shabbat_hours):
    additional_hours = max(hours - 8, 0)
    if shabbat_hours:
        additional_hours = 0
    base_sum = 35 if isPik else 30
    shabbat_sum = 45
    payment_for_regular_hours = (hours - additional_hours - shabbat_hours) * base_sum + (shabbat_hours * shabbat_sum)
    payment_for_additional_hours = additional_hours * (base_sum * 1.25)
    base_salary = payment_for_regular_hours + payment_for_additional_hours
    return math.ceil(base_salary)


def calculate_completion(base_salary, tip, base_completion):
    completion = base_completion if base_salary < tip + base_completion else base_salary - tip
    return completion


def calculate_average(ws, columns_count):
    current_row = find_first_empty(ws)
    sums = [0] * columns_count
    counter = 0
    for i in range(2, current_row):
        if ws.cell(i, 10).value != "pik":
            counter += 1
            for j in range(2, columns_count + 1):
                sums[j - 1] += ws.cell(i, j).value
    for i in range(len(sums)):
        sums[i] = sums[i] / counter if counter else sums[i]
    sums[0] = "Average"
    for i, num in enumerate(sums):
        ws.cell(current_row + 1, i + 1, '').font = Font(bold=False)
        ws.cell(current_row + 2, i + 1, num).font = Font(bold=True)


# def calculate_average2(ws, columns_count):
#     current_row = find_first_empty(ws)
#     sums = [0] * columns_count
#     for i in range(2, current_row):
#         for j in range(2, columns_count + 1):
#             if ws.cell(i, j) != 0:
#                 sums[j-1] = ws.cell[i, j] / current_row


# "%.2f" %

def update_sum(ws, columns_count):
    current_row = find_first_empty(ws)
    sums = [0] * columns_count
    for i in range(2, current_row):
        for j in range(2, columns_count + 1):
            if ws.cell(i, j).value != '':
                sums[j - 1] += float(ws.cell(i, j).value)
    sums[0] = 'sums'
    for i, sum in enumerate(sums):
        ws.cell(current_row + 3, i + 1, '').font = Font(bold=False)
        ws.cell(current_row + 4, i + 1, sum).font = Font(bold=True)


def update_worker(ws, date, hours, tip, tip_percentage, shabbat_hours, isPik, base_completion=100, overide=True):
    additional_hours = max(hours - 8, 0)
    base_salary = calculate_base_salary(hours, isPik, shabbat_hours)
    completion = calculate_completion(base_salary, tip, base_completion)
    overall = tip + completion
    worker_details = [date, hours, base_salary, tip, overall, completion, additional_hours, tip_percentage,
                      shabbat_hours, "pik" if isPik else ""]
    empty_row = find_first_empty(ws, date if overide else '')
    for column, detail in enumerate(worker_details):
        ws.cell(empty_row, column + 1, detail)
    update_sum(ws, 9)
    calculate_average(ws, 9)


def update_income(ws, date, x, credit, tip):
    overall_income = x + tip
    cash = x - credit
    income_details = [date, x, credit, cash, tip, overall_income]
    current_row = find_first_empty(ws, date)
    for column, detail in enumerate(income_details):
        ws.cell(current_row, column + 1, detail)
    update_sum(ws, 6)
    tip_percentage = float("%.2f" % (int(tip) / x * 100))
    print(tip_percentage)
    return tip_percentage

    # calculate_tip_average(ws, 6)


def new_workbook(workers_list, month, year):
    wb = op.Workbook()
    for worker in workers_list:
        add_sheet(wb, worker, categories)
    add_sheet(wb, months[month - 1] + " 20" + str(year), income_list)
    return wb


def update_pik(date, base_completion= 100, overide=True):
    pick = input('Please enter pik worker name, hours (shabbat hours) :')
    pick_details = pick.split()
    shabbat_hours = float(pick_details[2]) if len(pick_details) == 3 else 0
    update_worker(wb[pick_details[0]], date, float(pick_details[1]), 0, 0, shabbat_hours, True, base_completion, overide)


def regular_shift_update(tip_percentage, date, status, overide=True):
    open_shift = input('Please enter' + status + ' shift worker name, hours, tip, (shabbat hours, base completion): ')
    open_shift_details = open_shift.split()
    shabbat_hours = float(open_shift_details[3]) if len(open_shift_details) >= 4 else 0
    base_completion = float(open_shift_details[4]) if len(open_shift_details) == 5 else 100
    update_worker(wb[open_shift_details[0]], date, float(open_shift_details[1]), float(open_shift_details[2]),
                  tip_percentage, shabbat_hours, False, base_completion, overide)


def income_update(date):
    date_list = date.split(".")
    month = date_list[1]
    year = date_list[2]
    correct_date = months[int(month) - 1]
    x = input('Hello noam! Please enter daily x: ')
    credit = input('Please enter credit amount: ')
    tip = input('Please enter tip amount:')
    tip_percentage = update_income(wb[correct_date + " 20" + year], date, float(x), float(credit), float(tip))
    return tip_percentage


def update_day(is_shabbat=False):
    """
    root function to update all details
    :param is_shabbat: True if it is saturday
    """
    date = input('Please enter date:')
    tip_percentage = income_update(date)
    if is_shabbat:
        regular_shift_update(tip_percentage, date, " Noon ")
    regular_shift_update(tip_percentage, date, " open ")
    regular_shift_update(tip_percentage, date, " close ")
    update_pik(date)
    # update_pik(date , overide=False)
    # wb.save("worker tracing.xlsx")


# wb = op.Workbook()
# add_worker(wb, "Michal", categories)
wb = op.load_workbook("Workers tracing.xlsx")


# add_worker(wb, "", categories)
# wb = new_workbook(worker_list, 9, 21)
# regular_shift_update(6000,"1.9.21", "open ")
# update_worker(wb['Nitai'], "2.9.21", 8.5, 219, 0, False)
# update_worker(wb['Omri'], "2.10.21", 7, 144, 0, False)
# update_worker(wb['Omri'], "1.9.21", 9, 300,)
# update_day(wb)
# update_worker(wb['Rinat'], "25.8.21", 4, 0, 0, True)
# update_income(wb['September 2021'], "26.8.21", 4195, 3520, 489)
# update_day(wb)
# income_update('1.9.21')
# calculate_average(wb['Alon'], 9)
update_day(is_shabbat=True)

wb.save("Workers tracing.xlsx")
